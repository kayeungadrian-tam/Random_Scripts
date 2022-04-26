import numpy as np
import matplotlib.pyplot as plt
from scipy.integrate import odeint
import japanize_matplotlib

from openpyxl import Workbook
import pandas as pd

import os.path
import datetime
import string

# define model
def vehicle(v,t,u,load):
    # inputs
    #  v    = vehicle velocity (m/s)
    #  t    = time (sec)
    #  u    = gas pedal position (-50% to 100%)
    #  load = passenger load + cargo (kg)
    Cd = 0.24    # drag coefficient
    rho = 1.225  # air density (kg/m^3)
    A = 5.0      # cross-sectional area (m^2)
    Fp = 30      # thrust parameter (N/%pedal)
    m = 500      # vehicle mass (kg)
    # calculate derivative of the velocity
    dv_dt = (1.0/(m+load)) * (Fp*u - 0.5*rho*Cd*A*v**2)
    return dv_dt

tf = 300.0                 # final time for simulation
nsteps = 301               # number of time steps
delta_t = tf/(nsteps-1)   # how long is each time step?
# delta_t = 0.1
ts = np.linspace(0,tf,nsteps) # linearly spaced time vector

step = np.zeros(nsteps) # u = valve % open
# passenger(s) + cargo load
load = 200.0 # kg
# velocity initial condition
v0 = 0.0
# for storing the results
vs = np.zeros(nsteps)
sps = np.zeros(nsteps)

sp = 10

AC = [0]

tgt_A = [0]

ubias = 0.
Kc = 1./1.2 * 2.5
tauI = 20.0
sum_int = 0.0
es = np.zeros(nsteps)
ies = np.zeros(nsteps)

# simulate with ODEINT
for i in range(nsteps-1):
    if i == 50:
        sp = 25
    if i == 100:
        sp = 5
    u = step[i]
    # clip inputs to -50% to 100%
    if u >= 100.0:
        u = 100.0
        sum_int = sum_int - error * delta_t
    if u <= -50.0:
        u = -50.0
        sum_int = sum_int - error * delta_t

    sps[i+1] = sp
    error = sp - v0
    es[i+1] = error
    sum_int = sum_int + error * delta_t
    u = ubias + Kc*error + Kc/tauI*sum_int
    
    step[i+1] = u
    a = vehicle(vs[i],0,u,load)
    AC.append(a)

    v = odeint(vehicle,v0,[0,delta_t],args=(u,load))
    v0 = v[-1]   # take the last value
    vs[i+1] = v0 # store the velocity for plotting

    tgt_A.append(np.clip(sp - v0,-2.5,2.5))
    


alphabet = list(string.ascii_uppercase)
wb = Workbook()
ws = wb.active
label = [
    ('時間','s'),
    ('車速','m/s'),
    ('加速度','m/s^2'),
    ('目標車速','m/s'),
    ('目標加速度','m/s^2')
]
for idx, (k,v) in enumerate(label):
    ws[f'{alphabet[idx]}1']=k
    ws[f'{alphabet[idx]}2']=v

for row in range(nsteps-1):
    ws[f'A{row+3}'] = ts[row]
    ws[f'B{row+3}'] = vs[row]
    ws[f'C{row+3}'] = float(AC[row])
    ws[f'D{row+3}'] = float(sps[row])
    ws[f'E{row+3}'] = float(tgt_A[row])

wb.save('acc/acc.xlsx')

excel = pd.read_excel('acc/acc.xlsx', index_col=0)
excel.to_csv('acc/acc.csv', encoding='SHIFT-JIS')

from keras.models import Sequential
from keras.layers.core import Dense, Activation, Dropout
from keras.callbacks import EarlyStopping

data = pd.read_csv('acc/acc.csv',encoding="SHIFT-JIS", delimiter=',', skiprows=1)
header_column = pd.read_csv('acc/acc.csv',encoding="SHIFT-JIS", nrows=1)
header = [name for name in header_column]
data.columns = header

y = data['目標加速度'].values


columns_label = [
    '車速',
    # '加速度',
    '目標車速',
    # '目標加速度',
]

x = data[columns_label].values

model = Sequential()
model.add(Dense(12, input_dim=2, activation='relu'))
model.add(Dense(10, activation='relu'))
model.add(Dense(1, activation='linear'))

model.compile(loss='mse', optimizer='adam', metrics=['accuracy'])


es = EarlyStopping(monitor='loss', mode='min', verbose=1, patience=50)
history = model.fit(x, y, epochs=1000, batch_size=10, verbose=100, callbacks=[es]) 
model.save('acc/acc.h5')


FIG = plt.figure()
AX = FIG.add_subplot(111)
AX.plot(history.history["loss"], 'o-', ms=0.5, lw=0.7)
AX.set_title('model loss')
AX.set_xlabel('epoch')
AX.set_ylabel('loss')
AX.grid()

FIG = plt.figure()
AX = FIG.add_subplot(111)
AX.plot(history.history["accuracy"], 'o-', ms=0.5, lw=0.7)
AX.set_title('model accuracy')
AX.set_xlabel('epoch')
AX.set_ylabel('accuracy')
AX.grid()


plt.show()



# plot results
fig = plt.figure(figsize=(14,8))
ax1 = fig.add_subplot(311)
ax1.plot(ts,vs,'b-')
ax1.plot(ts,sps,'k--',lw=1)
ax1.set_ylabel('車速 [m/s]')
ax1.legend(['観測値','目標車速'],loc=2, fontsize=8)
plt.grid()

ax2 = fig.add_subplot(312)
ax2.plot(ts,AC,'r--',lw=1)
ax2.set_ylabel('加速度 [m/$s^2$]')
plt.grid()

ax3 = fig.add_subplot(313)
ax3.plot(ts,step,'g--',lw=1)
ax3.set_ylabel('アクセル開度 [%]')
ax3.set_xlabel('走行時間 [s]')
plt.grid()
plt.tight_layout()

fig.savefig('acc/acc.png')
# plt.show()