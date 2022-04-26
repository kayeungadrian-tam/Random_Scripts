import numpy as np
import matplotlib.pyplot as plt
from scipy.integrate import odeint
import japanize_matplotlib

from openpyxl import Workbook
import pandas as pd

import os.path
import datetime
import string

class Vehicle():
    def __init__(self):
        pass

    def update(self,v,t,u,load):
        Cd = 0.24    # drag coefficient
        rho = 1.225  # air density (kg/m^3)
        A = 5.0      # cross-sectional area (m^2)
        Fp = 30      # thrust parameter (N/%pedal)
        m = 500      # vehicle mass (kg)
        # calculate derivative of the velocity
        dv_dt = (1.0/(m+load)) * (Fp*u - 0.5*rho*Cd*A*v**2)
        return dv_dt

class PI_control():
    def __init__(self):
        self.ubias = 0
        self.Kc = 1./1.2 * 2.5
        self.tauI = 20.0
        self.sum_int = 0.0

    def update(self, set_point, v0):
        self.error = set_point - v0
        self.sum_int += self.error*delta_t
        u = self.ubias + self.Kc*self.error + self.Kc/self.tauI*self.sum_int
        return u

    def reset(self):
        self.sum_int = self.sum_int - self.error*delta_t

class Network():
    def __init__(self):
        from keras.models import Sequential
        from keras.layers.core import Dense, Activation, Dropout
        from keras.callbacks import EarlyStopping

        model = Sequential()
        model.add(Dense(256, input_dim=2, activation='relu'))
        model.add(Dense(128, activation='relu'))
        model.add(Dense(1, activation='linear'))

        model.compile(loss='mse', optimizer='adam', metrics=['accuracy'])

        self.model = model

    def load_data(self, file_path):
        data = pd.read_csv(file_path,encoding="SHIFT-JIS", delimiter=',', skiprows=1)
        header_column = pd.read_csv(file_path,encoding="SHIFT-JIS", nrows=1)
        header = [name for name in header_column]
        data.columns = header

        y = data['目標加速度'].values

        columns_label = [
            '車速',
            '加速度',
            # '目標車速',
            # '目標加速度',
        ]

        x = data[columns_label].values

        return x, y

    def fit(self, x, y):
        self.model.fit(x, y, epochs=500, batch_size=10, verbose=100)
        return self.model

    def save(self, path):
        self.model.save(path)

    def load(self,path):
        pass


def sampling(nsteps, model):

    def write_csv():
        alphabet = list(string.ascii_uppercase)
        wb = Workbook()
        ws = wb.active
        label = [
            ('時間','s'),
            ('車速','m/s'),
            ('加速度','m/s^2'),
            ('目標車速','m/s'),
            ('目標加速度','m/s^2')
        ]
        for idx, (k,v) in enumerate(label):
            ws[f'{alphabet[idx]}1']=k
            ws[f'{alphabet[idx]}2']=v

        for row in range(nsteps-1):
            ws[f'A{row+3}'] = ts[row]
            ws[f'B{row+3}'] = vs[row]
            ws[f'C{row+3}'] = float(AC[row])
            ws[f'D{row+3}'] = float(sps[row])
            ws[f'E{row+3}'] = float(tgt_A[row])

        wb.save('acc/acc.xlsx')

        excel = pd.read_excel('acc/acc.xlsx', index_col=0)
        excel.to_csv('acc/acc.csv', encoding='SHIFT-JIS')

    step = np.zeros(nsteps)
    vs = np.zeros(nsteps)
    sps = np.zeros(nsteps)
    AC = np.zeros(nsteps)
    tgt_A = np.zeros(nsteps)

    pred = np.zeros(nsteps)
    v_pred = np.zeros(nsteps)


    load = 200

    v0 = 0.0
    car = Vehicle()
    pid = PI_control()

    sp = 10
    state = np.zeros((1,2))
    for i in range(nsteps-1):
        if i == 50:
            sp = 25
        if i == 100:
            sp = 5
        u = step[i]
        if u >= 100.0:
            u = 100.0
            pid.reset()
        if u <= -50.0:
            u = -50.0
            pid.reset()


        sps[i+1] = sp
       
        u = pid.update(sp,v0)
        step[i+1] = u

        a = car.update(vs[i],0,u,load)
        AC[i+1] = a

        v = odeint(car.update,v0,[0,delta_t],args=(u,load))
        v0 = v[-1]   

        state[:,0] = v[-1]
        state[:,1] = a
 

        pred[i+1] = float(model.predict((state)))

        v_pred[i+1] = v_pred[i]+pred[i]*delta_t

        vs[i+1] = v0 
        # tgt_A[i+1] = (np.clip(sps[i+1] - sps[i],-10,10))
        tgt_A[i+1] = sps[i+1] - sps[i]




        


    plt.plot(ts,vs)
    plt.plot(ts,sps)
    plt.plot(ts,v_pred)
    # plt.plot(ts,AC)
    # plt.plot(ts,pred)
    plt.grid()

    write_csv()

    plt.show()


if __name__ == '__main__':
    tf = 300.0                
    nsteps = 301               
    delta_t = tf/(nsteps-1)   
    ts = np.linspace(0,tf,nsteps) 

    nn = Network()

    x, y = nn.load_data('acc/acc.csv')

    model = nn.fit(x,y)

    nn.save('acc/test.h5')

    sampling(nsteps=nsteps, model=model)   

