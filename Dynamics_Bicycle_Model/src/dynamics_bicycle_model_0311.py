import numpy as np
import matplotlib.pyplot as plt
import yaml
import japanize_matplotlib

from openpyxl import Workbook
import openpyxl
import pandas as pd
import string

import os.path
from os import path

import datetime
from tqdm import tqdm


with open('config.yaml') as file:
    params = yaml.load(file, Loader=yaml.FullLoader)

    dt =    params['dt']
    l =     params['l']
    lf =    params['lf']   
    lr = l - lf

    Cf =    params['Cf']
    Cr =    params['Cr']
    Iz =    params['Iz']
    m =     params['m']

    tp =    params['tp']

    v =     params['v']/3.6

    A = (-m*(lf*Cf - lr*Cr))/(2*l**2*Cr*Cf)

    radius = params['radius']
    straight_len = params['straight_len']

    x = params['x_init']
    y = params['y_init']

    yaw = params['yaw']

    real_steering = (17)*(1/250)*(l/(1+A*v**2))


class Course():
    def __init__(self):
        theta = np.linspace(-np.pi/2, np.pi/4, 40000)
        r = np.sqrt(radius**2)
        x1 = r*np.cos(theta) + straight_len
        x2 = r*np.sin(theta) + (radius+y)
        result = np.hstack((x1.reshape(-1,1),x2.reshape(-1,1)))

        xs = np.linspace(0,straight_len,1000)
        ys = np.array([10]*1000)
        lower = np.hstack((xs.reshape(-1,1),ys.reshape(-1,1)))
        final = np.concatenate((lower,result))

        self.track_all = final

class Vehicle():
    def __init__(self, x, y, v, yaw, yaw_rate, beta):
        course = Course()
        self.full_track = course.track_all

        self.count = 0 
        self.x = x
        self.y = y

        self.v = v

        self.yaw = yaw
        self.yaw_rate = yaw_rate
        self.beta = beta 

    def predict(self):
        self.gamma = (self.delta*self.v)/((1+A*self.v**2)*l)

        if self.delta == 0:
            self.sin_pre = 1
            self.cos_pre = 0
        else:
            self.sin_pre = np.sin(self.gamma*tp)/self.gamma
            self.cos_pre = (1 - np.cos(self.gamma*tp))/self.gamma
        
        self.xx_dot = self.v*self.sin_pre
        self.yy_dot = self.v*self.cos_pre

        self.x_pred = self.x + dt*(np.cos(self.yaw)*self.xx_dot - np.sin(self.yaw)*self.yy_dot)
        self.y_pred = self.y + dt*(np.sin(self.yaw)*self.xx_dot + np.cos(self.yaw)*self.yy_dot)

        self.curve = (self.beta_dot+self.yaw_rate)/self.v

        def rotate(P, theta):
            c, s = np.cos(theta), np.sin(theta)
            R = np.array(([c, s], [-s, c]), dtype=np.float)
            return P.dot(R)

        track_global = self.full_track
        # track_local = rotate(track_global - (self.x + np.cos(self.yaw)*lf, self.y + np.sin(self.yaw)*lf), -self.yaw)
        track_local = rotate(track_global - (self.x_pred, self.y_pred), -self.yaw)
        
        distances = np.linalg.norm(track_local - (0,0), axis=1)
        min_distance = np.min(distances)
        min_index = np.argmin(distances)

        self.heading_track = np.arctan((track_local[min_index][1]-track_local[min_index-1][1])/(track_local[min_index][0]-track_local[min_index-1][0]))
        self.yaw_ss = m*self.v*self.heading_track/(Cf*(1+(lf/lr)))

        self.H = np.sqrt((self.x_pred-self.x)**2+(self.y_pred-self.y)**2)

        check = (0 - track_local[min_index-1][0])*(track_local[min_index][1]-track_local[min_index-1][1]) \
            - (0 -track_local[min_index-1][1])*(track_local[min_index][0] - track_local[min_index-1][0])

        if check < 0:
            sign = -1
        elif check > 0:
            sign = 1
        else:
            sign = 0

        feedback = sign*min_distance
        self.cte = sign*min_distance
        return feedback

    def update(self, delta):

        def normalize_angle(angle):
            if angle > np.pi:
                angle = angle - 2 * np.pi
            if angle < -np.pi:
                angle = angle + 2 * np.pi
            return angle
        
        self.yaw = normalize_angle(self.yaw)
        self.beta = normalize_angle(self.beta)
        
        self.count += 1
        self.delta = np.clip(delta, -np.pi/4, np.pi/4)

        self.gamma = (self.delta*self.v)/((1+A*self.v**2)*l)

        if self.delta == 0:
            self.sin1 = 1
            self.cos1 = 0
        else:
            self.sin1 = np.sin(self.gamma)/self.gamma 
            self.cos1 = (1-np.cos(self.gamma))/self.gamma

        self.x_dot = (self.v)*(self.sin1)
        self.y_dot = (self.v)*(self.cos1)

        self.x += dt*(np.cos(self.yaw)*self.x_dot - np.sin(self.yaw)*self.y_dot)
        self.y += dt*(np.sin(self.yaw)*self.x_dot + np.cos(self.yaw)*self.y_dot )

        self.beta_dot = (-(Cr+Cf)/(m*self.v))*self.beta + (((Cr*lr-Cf*lf)/(m*self.v**2))-1)*self.yaw_rate + Cf*self.delta/(m*self.v)
        self.yaw_rate_dot = (Cr*lr - Cf*lf)*self.beta/Iz - (Cr*lr**2 + Cf*lf**2)*self.yaw_rate/(Iz*self.v) + Cf*lf*self.delta/Iz

        self.beta = self.beta + self.beta_dot*dt
        self.yaw_rate = self.yaw_rate + self.yaw_rate_dot*dt

        self.yaw =  self.yaw + self.yaw_rate*dt

class Stanley():
    def __init__(self, K, Kdy, Kds):
        self.K = K
        self.Kdy = Kdy
        self.Kds = Kds
        self.reset()

    def reset(self):
        self.stanley_value = 0

    def update(self, error, phi, phi_ss, v,  r_meas, r_traj, delta):
        new_stanley_value = (phi - phi_ss) +\
                                np.arctan(self.K*error/v) + \
                                    self.Kdy*(r_meas-r_traj) + \
                                        self.Kds*(self.stanley_value - delta)
        self.stanley_value = new_stanley_value

class PID():
    def __init__(self, P=0.2, I=0.0, D=0., deltatime=0.01, windup_guard=20.0):
        self.Kp=P
        self.Ki=I
        self.Kd=D
        self.deltatime=deltatime
        self.windup_guard=windup_guard
        self.reset()

    def reset(self):
        self.PTerm = 0.0
        self.ITerm = 0.0
        self.DTerm = 0.0
        self.last_error = 0.0

    def update(self, feedback_value, target_value):
        error = target_value - feedback_value
        delta_error = error - self.last_error
        self.PTerm = self.Kp * error
        self.ITerm += error * self.deltatime
        self.ITerm = np.clip(self.ITerm, -self.windup_guard, self.windup_guard)

        self.DTerm = delta_error / self.deltatime
        self.last_error = error

        output = self.PTerm + (self.Ki * self.ITerm) + (self.Kd * self.DTerm)
        return output

def sampling(total_steps, out_path, apply_nn=False):
    NN_result, Feedback_list = [], []
    if apply_nn:
        from keras.models import Sequential
        from keras.layers.core import Dense, Activation, Dropout
        from keras.models import load_model

        model = load_model(f'{out_path}/saved_model/model.h5')
        out_path = f'{out_path}/Sim_data'
        if not path.exists(out_path):
            os.mkdir(out_path)

    else:
        out_path = f'{out_path}/Sample_data'
        if not path.exists(out_path):
            os.mkdir(out_path)
        nn_pred= 0
        
    def write_csv(vector, total_steps):
        wb = Workbook()
        ws = wb.active
        ws.title = 'log_data'
        alphabet = list(string.ascii_uppercase)
        columns_label = [
            ('走行時間','[s]'),
            ('x_座標','[m]'),
            ('y_座標','[m]'),
            ('横車速','[km/h]'),
            ('縦車速','[km/h]'),
            ('ヨー角','[rad]'),
            ('ヨーレイト','[rad]'),
            ('横滑り角','[rad]'),
            ('横滑り角レイト','[rad]'),
            ('タイヤ角','[rad]'),
            ('cte','[m]'),
            ('曲率_車両','[1/m]'),
            ('曲率_コース','[1/m]'),
            ('ヨ―角_コース','[rad]')
        ]

        for idx, (k,value) in enumerate(columns_label):
            ws[f'{alphabet[idx]}1'] = k
            ws[f'{alphabet[idx]}2'] = value

        if apply_nn:
            ws['P1'] = 'NN_pred'
            ws['P2'] = '[rad]'
            ws['Q1'] = 'PID_value'
            ws['Q2'] = '[rad]'


        for step in range(total_steps-1):
            # print(vector)
            for sub_len in range(vector.shape[1]):
                ws[f'{alphabet[sub_len]}{step+3}'] = vector[step][sub_len]

                if apply_nn:
                    ws[f'P{step+3}'] = NN_result[step]
                    ws[f'Q{step+3}'] = Feedback_list[step]

        wb.save(f'{out_path}/00_log-{datetime.date.today()}.xlsx')
        excel = pd.read_excel(f'{out_path}/00_log-{datetime.date.today()}.xlsx', index_col=0)
        excel.to_csv(f'{out_path}/00_log-{datetime.date.today()}.csv',encoding="SHIFT-JIS")

    def plot_trajectory(vector, close_up=False):
        fig = plt.figure()
        ax = fig.add_subplot(111)
        fig.suptitle('走行軌跡')
        ax.set_xlabel('x 座標 [m]')
        ax.set_ylabel('y 座標 [m]')
        ax.plot(car.full_track[:,0], car.full_track[:,1], 'o-', c='r',ms=0.5, alpha=0.4, label='course')
        ax.plot(vector[:,1], vector[:,2], ls='dashed',c='k',lw=1.5, ms=0.7, label='path')       
        ax.legend(fontsize=8)
        ax.grid()
        if close_up:
            ax.set_xlim(np.max(vector[:,1])-0.1,np.max(vector[:,1])+0.1)
            ax.set_ylim(np.max(vector[:,2])-0.1,np.max(vector[:,2])+0.1)
            fig.tight_layout()
            fig.savefig(out_path+'/01_走行軌跡_拡大')
        else:
            fig.tight_layout()
            fig.savefig(out_path+'/01_走行軌跡')
        
        fig2, axs = plt.subplots(2,1, sharex=True, figsize=(12,6))
        fig2.suptitle('走行軌跡 X, Y')
        axs = axs.flat
        axs[0].plot(vector[:,0],vector[:,1])
        axs[0].set_ylabel('x 座標 [m]')
        axs[1].plot(vector[:,0],vector[:,2])
        axs[1].set_xlabel('走行時間 [s]')
        axs[1].set_ylabel('y 座標 [m]')
        axs[0].grid()
        axs[1].grid()
        fig2.savefig(out_path+'/01_走行軌跡_xy')

    def plot_data(vector, index):
        columns_label = [
            ('走行時間','[s]'),
            ('x_座標','[m]'),
            ('y_座標','[m]'),
            ('横車速','[km/h]'),
            ('縦車速','[km/h]'),
            ('ヨー角','[rad]'),
            ('ヨーレイト','[rad]'),
            ('横滑り角','[rad]'),
            ('横滑り角レイト','[rad]'),
            ('タイヤ角','[rad]'),
            ('cte','[m]'),
            ('曲率_車両','[1/m]'),
            ('曲率_コース','[1/m]'),
            ('ヨ―角_コース','[rad]')
        ]

        fig = plt.figure(figsize=(12,6))
        ax = fig.add_subplot(111)

        ax.set_xlabel('走行時間 [s]')
        ax.set_ylabel(f'{columns_label[index][1]}')

        if index == 11:
            fig.suptitle('曲率')
            ax.plot(vector[:,0],vector[:,11], label='車両')
            ax.plot(vector[:,0],vector[:,12], label='コース')
            ax.legend()
            ax.grid()
            ax.set_ylim(-0.002,0.005)
            fig.tight_layout()
            fig.savefig(f'{out_path}/{str(index).zfill(2)}_曲率')
        elif index == 12:
            pass
        elif index == 9:
            fig.suptitle('タイヤ角 (比率: 1:17)')
            ax.plot(vector[:,0], vector[:,9], label='実測値')
            ax.axhline(y=real_steering, label='理想: l/$\\rho$(1+A$v^2$)', c='r')
            ax.set_xlabel('走行時間 [s]')
            ax.set_ylabel('[rad]')
            ax.set_ylim(-0.02,0.07)
            ax.legend()
            ax.grid()
            fig.tight_layout()
            fig.savefig(f'{out_path}/{str(index).zfill(2)}_{columns_label[index][0]}')
        else:
            fig.suptitle(columns_label[index][0])
            ax.plot(vector[:,0], vector[:,index])
        
            ax.grid()
            fig.tight_layout()
            fig.savefig(f'{out_path}/{str(index).zfill(2)}_{columns_label[index][0]}')

    def update_state_vector(idx,s):
        state_vector[idx][0] = idx/100+0.01  # 走行時間
        state_vector[idx][1] = s.x      # X 座標
        state_vector[idx][2] = s.y      # Y 座標
        state_vector[idx][3] = s.v*3.6      # 横車速
        state_vector[idx][4] = 3.6*s.v*(s.beta+s.yaw) # 縦車速
        state_vector[idx][5] = s.yaw    # ヨー角
        state_vector[idx][6] = s.yaw_rate   # ヨーレイト
        state_vector[idx][7] = s.beta       # 横滑り角
        state_vector[idx][8] = s.beta_dot   # 横滑り角例と
        state_vector[idx][9] = s.delta  # タイヤ角
        state_vector[idx][10] = s.cte   # クロス　トラック　エラー
        state_vector[idx][11] = s.curve # 曲率（車両）
        state_vector[idx][12] = (0 if s.x < 50 else 0.004) # 曲率（コース）
        state_vector[idx][13] =  s.heading_track # ヨー角（コース）

    car = Vehicle(
        x = x,
        y = y,
        v = v,
        yaw = yaw,
        yaw_rate = 0,
        beta = 0.0
    )

    pid = PID(
        P=1.5,
        I=0.7, 
        D=0.0002
        )

    feedback = 0.0
    state_vector = np.zeros((total_steps,14))


    print('Running simulation...')
    for i in tqdm(range(total_steps), leave=True):
        car.update(delta = feedback)
        cross_track_error = car.predict()
        update_state_vector(i,car)
        
        if apply_nn:
            state = np.zeros((1,2))

            state[:,0] = state_vector[i][3]/50
            # state[:,1] = state_vector[i][11]
            state[:,1] = state_vector[i][12]

            nn_pred = float(model.predict(state))

            NN_result.append(nn_pred)

        pid_value = pid.update(cross_track_error,0)
        Feedback_list.append(pid_value)

        # if abs(pid_value) < 0.005:
        #     feedback -=  pid_value
        # else:
        feedback -=  pid_value + nn_pred

        update_state_vector(i,car)

    # if apply_nn:
    fig_nn = plt.figure(figsize=(14,7))
    ax_nn = fig_nn.add_subplot(111)
    ax_nn.set_title('Feedback & NN result')
    ax_nn.plot(NN_result, label='NN')
    ax_nn.plot(Feedback_list, label='FB')
    # ax_nn.plot(np.array(NN_result)+np.array(Feedback_list), label='NN + FB')
    ax_nn.grid()
    ax_nn.legend()
    fig_nn.savefig(f'{out_path}/99_NN_result.png')

    print('Saving csv file...')
    write_csv(state_vector, total_steps)
    print('Saving log data plots...')
    plot_trajectory(state_vector)
    plot_trajectory(state_vector, close_up=True)
    for k in range(3,14):
        plot_data(state_vector, k)
    print('Successful.')

def my_nomralize(dataframe):
    array1 = dataframe[:,0]
    array2 = dataframe[:,1]
    dataframe[:,0] = array1/array1.max()
    # dataframe[:,1] = (array2 - array2.min())/(array2.max()-array2.min())
    dataframe[:,1] = dataframe[:,1]/50
    return dataframe

def training(epoch=1000):
    model_dir = f'{out_path}/saved_model'
    if not path.exists(model_dir):
        os.mkdir(model_dir)
    from keras.models import Sequential
    from keras.layers.core import Dense, Activation, Dropout
    data = pd.read_csv(f'{out_path}/sample_data/00_log-{datetime.date.today()}.csv',encoding="SHIFT-JIS", delimiter=',', skiprows=1)
    header_column = pd.read_csv(f'{out_path}/sample_data/00_log-{datetime.date.today()}.csv',encoding="SHIFT-JIS", nrows=1)
    header = [name for name in header_column]
    data.columns = header

    Y = data['タイヤ角'].values/50

    columns_label = [
        # '走行時間',
        # 'x_座標',
        # 'y_座標',
        '横車速',
        # '縦車速',
        # 'ヨー角',
        # 'ヨーレイト',
        # '横滑り角',
        # '横滑り角レイト',
        # 'タイヤ角',
        # 'cte',
        # '曲率_車両',
        '曲率_コース',
        # 'ヨ―角_コース',
    ]

    X = data[columns_label].values
    
    X = my_nomralize(X)


    model = Sequential()
    model.add(Dense(12, input_dim=2, activation='relu'))
    model.add(Dense(10, activation='relu'))
    model.add(Dense(1, activation='linear'))

    model.compile(loss='mse', optimizer='adam', metrics=['accuracy'])
    model.save(f'{model_dir}/model.h5')

    history = model.fit(X, Y, epochs=epoch, batch_size=10, verbose=100)
    
    FIG = plt.figure()
    AX = FIG.add_subplot(111)
    AX.plot(history.history["loss"], 'o-', ms=0.5, lw=0.7)
    AX.set_title('model loss')
    AX.set_xlabel('epoch')
    AX.set_ylabel('loss')
    AX.grid()
    FIG.savefig(f'{model_dir}/loss_full.png')
    AX.set_ylim(0,15*np.min(history.history["loss"]))
    FIG.savefig(f'{model_dir}/loss_closeup.png')

Training = True

if __name__=='__main__':

    now = datetime.datetime.now()
    title = now.strftime('%m%d')

    out_dir = '../output/'
    filename = f'00_sim-log-{title}'

    out_path = (os.path.join(out_dir,filename))

    if not path.exists(out_path):
        os.mkdir(out_path)

    # sampling(2000,out_path=out_path)

    if Training:
        training(epoch=1000)
        
    sampling(2000, out_path=out_path, apply_nn=True)

        